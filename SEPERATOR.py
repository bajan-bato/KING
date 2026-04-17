#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_CSV = "data/G2_ planiranje.csv"
OUTPUT_EXCEL = "out/G2_ planiranje.xlsx"

QTY_RE = re.compile(r'(\d+)\s*[x×]?\s*kom\.?', re.IGNORECASE)

LIGHT_BLUE = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
LIGHT_BEIGE = PatternFill(start_color="FDF5E6", end_color="FDF5E6", fill_type="solid")
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

WARN_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
WARN_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PLACEHOLDER_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def extract_quantity_and_clean(address):
    if pd.isna(address) or address == '':
        return address, None
    match = QTY_RE.search(str(address))
    if match:
        qty = int(match.group(1))
        cleaned = QTY_RE.sub('', str(address)).strip()
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()
        return cleaned, qty
    return address, None

def process_row(row, elo_col, align_cols, output_row_start):
    elo_val = row.get(elo_col, '')
    if pd.isna(elo_val) or str(elo_val).strip() == '':
        new_row = row.to_dict()
        for k, v in new_row.items():
            if isinstance(v, list):
                new_row[k] = ', '.join(str(i) for i in v) if v else ''
        return [new_row], [], set()

    elo_str = str(elo_val).strip()
    elo_list = [e.strip() for e in elo_str.split('\n') if e.strip()]
    if not elo_list:
        new_row = row.to_dict()
        for k, v in new_row.items():
            if isinstance(v, list):
                new_row[k] = ', '.join(str(i) for i in v) if v else ''
        return [new_row], [], set()

    col_lists = {}
    for col in align_cols:
        val = row.get(col, '')
        if pd.isna(val):
            lst = ['']
        else:
            lines = [line.strip() for line in str(val).split('\n') if line.strip()]
            lst = lines if lines else ['']
        col_lists[col] = lst

    addr_list = col_lists.get('Adresa/e isporuke', [''])
    kontakt_list = col_lists.get('Kontakt broj', [''])
    ime_list = col_lists.get('Ime i prezime', [''])
    email_list = col_lists.get('e-mail', [''])

    need_explode = (len(elo_list) > 1) or (len(addr_list) > 1)

    if not need_explode:
        new_row = row.to_dict()
        new_row[elo_col] = elo_list[0] if elo_list else ''
        addr = addr_list[0] if addr_list else ''
        cleaned_addr, qty = extract_quantity_and_clean(addr)
        new_row['Adresa/e isporuke'] = cleaned_addr
        if qty is not None:
            new_row['Kom.'] = int(qty)
        new_row['Kontakt broj'] = '\n'.join(kontakt_list) if kontakt_list else ''
        new_row['Ime i prezime'] = '\n'.join(ime_list) if ime_list else ''
        new_row['e-mail'] = '\n'.join(email_list) if email_list else ''
        for k, v in new_row.items():
            if isinstance(v, list):
                new_row[k] = ', '.join(str(i) for i in v) if v else ''
        return [new_row], [], set()

    max_len = max(len(elo_list), len(addr_list))
    mismatches = None
    if len(elo_list) != len(addr_list):
        mismatches = f"Broj ELO={len(elo_list)}, Adresa={len(addr_list)}"

    id_col = None
    for col in row.index:
        if col.lower() == 'id':
            id_col = col
            break
    id_val = row.get(id_col, '') if id_col else ''

    new_rows = []
    warnings = []
    placeholder_indices = set()

    for i in range(max_len):
        new_row = row.to_dict()
        el = elo_list[i] if i < len(elo_list) else ''
        new_row[elo_col] = el

        addr = addr_list[i] if i < len(addr_list) else ''
        cleaned_addr, qty = extract_quantity_and_clean(addr)
        new_row['Adresa/e isporuke'] = cleaned_addr
        if qty is not None:
            new_row['Kom.'] = int(qty)
        else:
            new_row['Kom.'] = 1
            if addr.strip() != '':
                output_idx = output_row_start + len(new_rows)
                warnings.append((output_idx, el, id_val, f"Nedostaje količina u adresi: '{addr}'"))
                placeholder_indices.add(output_idx)

        new_row['Kontakt broj'] = kontakt_list[i] if i < len(kontakt_list) else (kontakt_list[0] if kontakt_list else '')
        new_row['Ime i prezime'] = ime_list[i] if i < len(ime_list) else (ime_list[0] if ime_list else '')
        new_row['e-mail'] = email_list[i] if i < len(email_list) else (email_list[0] if email_list else '')

        for k, v in new_row.items():
            if isinstance(v, list):
                new_row[k] = ', '.join(str(i) for i in v) if v else ''
        new_rows.append(new_row)

        if mismatches:
            output_idx = output_row_start + len(new_rows) - 1
            warnings.append((output_idx, el, id_val, f"Nepodudaranje broja stavki: {mismatches}"))

    return new_rows, warnings, placeholder_indices

def add_sums_to_title_rows(df, elo_col):
    """
    For each title row (empty ELO), sum the Kom. values of all following data rows
    until the next title row (or end of DataFrame). Write the sum into the title row's Kom. column.
    """
    kom_col = 'Kom.'
    if kom_col not in df.columns:
        df[kom_col] = 1
    df[kom_col] = pd.to_numeric(df[kom_col], errors='coerce').fillna(1).astype(int)
    
    # Find indices of title rows
    title_indices = []
    for idx, row in df.iterrows():
        elo_val = row.get(elo_col, '')
        if pd.isna(elo_val) or str(elo_val).strip() == '':
            title_indices.append(idx)
    
    # Process each block between title rows
    for i, title_idx in enumerate(title_indices):
        # Determine the end of the block: next title index or end of df
        next_title_idx = title_indices[i+1] if i+1 < len(title_indices) else len(df)
        # Data rows are those strictly after title_idx and before next_title_idx
        data_indices = range(title_idx + 1, next_title_idx)
        if not data_indices:
            continue
        block_sum = df.loc[data_indices, kom_col].sum()
        # Write the sum into the title row's Kom. cell
        df.at[title_idx, kom_col] = int(block_sum)
    
    return df

def main():
    useful_cols = ['PPZ', 'ELO', 'ID', 'Tip opreme ', 'Ustanove', 'Kom.', 'Lokacija', 
                   'Ime i prezime', 'Kontakt broj', 'e-mail', 'Matična adresa', 
                   'Adresa/e isporuke', 'OIB', 'Isporuka na lokaciju ', 
                   'Montaža\nstart', 'Montaža\nfinish', 'Instalacija ', 
                   'Održavanje ', 'Komentar']
    
    os.makedirs("out", exist_ok=True)

    df = pd.read_csv(INPUT_CSV, encoding='utf-8-sig', dtype=str, keep_default_na=False, usecols=useful_cols)
    print(f"Original rows: {len(df)}")
    original_cols = list(df.columns)

    elo_col = None
    for col in original_cols:
        if col.lower() == 'elo':
            elo_col = col
            break
    if elo_col is None:
        print("Error: 'ELO' column not found")
        return

    align_cols = ['Adresa/e isporuke', 'Kontakt broj', 'Ime i prezime', 'e-mail']
    for col in align_cols:
        if col not in df.columns:
            df[col] = ''

    all_rows = []
    all_warnings = []
    all_placeholders = set()

    for _, row in df.iterrows():
        out_rows, warns, placeholders = process_row(row, elo_col, align_cols, len(all_rows))
        all_rows.extend(out_rows)
        all_warnings.extend(warns)
        all_placeholders.update(placeholders)

    df_out = pd.DataFrame(all_rows)
    print(f"Output rows: {len(df_out)}")

    if 'Kom.' in df_out.columns:
        df_out['Kom.'] = pd.to_numeric(df_out['Kom.'], errors='coerce').fillna(1).astype(int)
    else:
        df_out['Kom.'] = 1

    # --- NEW: Add sums to title rows ---
    df_out = add_sums_to_title_rows(df_out, elo_col)
    title_rows_updated = df_out[df_out[elo_col].isna() | (df_out[elo_col].astype(str).str.strip() == '')]
    print(f"  Updated {len(title_rows_updated)} title rows with column sums.")

    final_cols = [col for col in original_cols if col in df_out.columns]
    for col in df_out.columns:
        if col not in final_cols and col != 'Kom.':
            final_cols.append(col)
    df_out = df_out[final_cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "Planiranje"

    for c, col in enumerate(final_cols, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.fill = LIGHT_BLUE
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r, row in df_out.iterrows():
        elo_val = row.get(elo_col, '')
        is_title = pd.isna(elo_val) or str(elo_val).strip() == ''
        fill = LIGHT_BLUE if is_title else LIGHT_BEIGE
        for c, col in enumerate(final_cols, 1):
            val = row.get(col, '')
            if isinstance(val, list):
                val = ', '.join(str(v) for v in val) if val else ''
            elif val is None or (isinstance(val, float) and pd.isna(val)):
                val = ''
            else:
                val = str(val)
            cell = ws.cell(row=r+2, column=c, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if col == 'Kom.' and r in all_placeholders:
                cell.fill = PLACEHOLDER_RED

    for c, col in enumerate(final_cols, 1):
        max_len = len(col)
        for r in range(2, len(df_out) + 2):
            cell_value = ws.cell(row=r, column=c).value
            if cell_value:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 40)

    if all_warnings:
        ws_warn = wb.create_sheet("Warnings")
        warn_headers = ["Redak izlaza", "EL", "ID", "Upozorenje"]
        for c, h in enumerate(warn_headers, 1):
            cell = ws_warn.cell(row=1, column=c, value=h)
            cell.fill = LIGHT_BLUE
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER

        for r, (row_idx, el, id_val, msg) in enumerate(all_warnings, 2):
            if "Nepodudaranje" in msg:
                fill_color = WARN_GREEN
            elif "Nedostaje količina" in msg:
                fill_color = WARN_YELLOW
            else:
                fill_color = WARN_RED
            for c, value in enumerate([row_idx + 2, el, id_val, msg], 1):
                cell = ws_warn.cell(row=r, column=c, value=value)
                cell.fill = fill_color
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for c, col in enumerate(warn_headers, 1):
            max_len = len(col)
            for r in range(2, len(all_warnings)+2):
                val = ws_warn.cell(row=r, column=c).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws_warn.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 40)

    wb.save(OUTPUT_EXCEL)
    print(f"Spremljeno: {OUTPUT_EXCEL}")
    if all_warnings:
        print(f"\n⚠️  {len(all_warnings)} upozorenja generirana (pogledajte list 'Warnings').")
    else:
        print("\n✅ Nema upozorenja.")
    if all_placeholders:
        print(f"ℹ️  {len(all_placeholders)} redova ima placeholder '1' u stupcu 'Kom.' (označeno crvenom bojom).")

if __name__ == "__main__":
    main()