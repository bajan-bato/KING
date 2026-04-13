#!/usr/bin/env python3
"""
Update Excel sheets (G1_ planiranje, G2_ planiranje) based on model2.csv.
For each Excel row, check which ELs (column C) have a matching CSV entry with a specific number.
Write a custom value into column B (full match or "Samo ..." partial).
Preserves Excel formatting.
Prints CSV entries that were not matched in Excel.
Usage:
    python update_excel_from_csv.py --override
"""

import os
import re
import csv
import shutil
import argparse
from pathlib import Path
from openpyxl import load_workbook

# ========== CONFIGURATION ==========
#EXCEL_PATH = "data/CARNET_G1_G2_plan_v0.15.xlsx"
EXCEL_PATH = "data/updated_excel.xlsx"
CSV_PATH = "output/model2.csv"
OUTPUT_DIR = "out"
OUTPUT_FILE = "updated_excel.xlsx"

# Which number to look for in the CSV (column "Number")
TARGET_NUMBER = "2"
# What value to write into column B when matched
VALUE_TO_WRITE = "2"
# ===================================

def load_csv_data(csv_path, target_number):
    """
    Load model2.csv and return:
        csv_matches: set of (group, el, id) that have the target_number.
        csv_all: dict {(group, el, id): row_info} for all rows with target number (for reporting).
    """
    matches = set()
    all_rows = {}  # store full row info for unmatched reporting
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                el = row.get('EL', '').strip()
                id_val = row.get('ID', '').strip()
                number = row.get('Number', '').strip()
                group_str = row.get('Group', '').strip()
                tip = row.get('Tip', '').strip()
                if not el or not id_val or not number or not group_str:
                    continue
                if group_str == "G1":
                    group = 1
                elif group_str == "G2":
                    group = 2
                else:
                    continue
                if number == target_number:
                    matches.add((group, el, id_val))
                    all_rows[(group, el, id_val)] = (el, id_val, group_str, tip, number)
    except Exception as e:
        print(f"Error reading CSV: {e}")
    return matches, all_rows

def strip_el_suffix(el):
    m = re.match(r'^([0-9]+)[a-z]?$', el)
    return m.group(1) if m else el

def el_matches(excel_el, csv_el):
    return excel_el == csv_el or strip_el_suffix(excel_el) == strip_el_suffix(csv_el)

def update_sheet(wb, sheet_name, group_num, csv_matches, all_csv_data, override, value_to_write, matched_excel_keys):
    """
    Update column B in the given sheet.
    Also add to matched_excel_keys set for every (group, el, id) that was successfully matched.
    """
    if sheet_name not in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' not found, skipping.")
        return
    ws = wb[sheet_name]
    max_row = ws.max_row
    updated_rows = 0

    for row_idx in range(1, max_row + 1):
        el_cell = ws.cell(row=row_idx, column=3)
        id_cell = ws.cell(row=row_idx, column=4)
        el_val = str(el_cell.value) if el_cell.value is not None else ""
        id_val = str(id_cell.value) if id_cell.value is not None else ""
        if not el_val.strip() or not id_val.strip():
            continue

        excel_els = [el.strip() for el in re.split(r'[\n\r]+', el_val) if el.strip()]
        if not excel_els:
            continue

        present = []
        for excel_el in excel_els:
            # Try exact match
            found = False
            if (group_num, excel_el, id_val) in csv_matches:
                found = True
                matched_excel_keys.add((group_num, excel_el, id_val))
            else:
                # Try stripped
                stripped = strip_el_suffix(excel_el)
                if stripped != excel_el and (group_num, stripped, id_val) in csv_matches:
                    found = True
                    matched_excel_keys.add((group_num, stripped, id_val))
            if found:
                present.append(excel_el)

        if not present:
            continue

        # Build new content
        if len(present) == len(excel_els):
            new_text = value_to_write
        else:
            new_text = f"Samo {', '.join(present)} {value_to_write}"

        b_cell = ws.cell(row=row_idx, column=2)
        current = b_cell.value if b_cell.value is not None else ""

        if override:
            b_cell.value = new_text
            updated_rows += 1
            print(f"  Row {row_idx}: overwrote -> '{new_text}'")
        else:
            if current:
                if new_text not in current.split('\n'):
                    b_cell.value = current + "\n" + new_text
                    print(f"  Row {row_idx}: appended -> '{new_text}'")
                else:
                    print(f"  Row {row_idx}: skipped (already has '{new_text}')")
            else:
                b_cell.value = new_text
                print(f"  Row {row_idx}: added -> '{new_text}'")
            updated_rows += 1

    print(f"  Updated {updated_rows} rows in sheet '{sheet_name}'.")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--override", action="store_true",
                        help="Overwrite existing content in column B (otherwise append)")
    args = parser.parse_args()

    # Load CSV data
    csv_matches, all_csv_rows = load_csv_data(CSV_PATH, TARGET_NUMBER)
    print(f"Loaded {len(csv_matches)} (group, el, id) entries with number '{TARGET_NUMBER}' from CSV.")

    if not csv_matches:
        print("No matching entries found. Exiting.")
        return

    # Prepare output
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, OUTPUT_FILE)
    shutil.copy2(EXCEL_PATH, output_path)
    print(f"Copied original Excel file to {output_path}")

    wb = load_workbook(output_path)
    matched_excel_keys = set()  # track which CSV keys were actually matched in Excel

    for group, sheet_name in [(1, "G1_ planiranje"), (2, "G2_ planiranje")]:
        print(f"\nProcessing sheet '{sheet_name}'...")
        update_sheet(wb, sheet_name, group, csv_matches, all_csv_rows, args.override, VALUE_TO_WRITE, matched_excel_keys)

    wb.save(output_path)
    print(f"\nSaved updated workbook to {output_path}")

    # Report unmatched CSV entries (those not found in any Excel sheet)
    unmatched_csv = csv_matches - matched_excel_keys
    if unmatched_csv:
        print("\n" + "="*60)
        print("CSV ENTRIES NOT MATCHED IN EXCEL (with target number):")
        for (group, el, id_val) in sorted(unmatched_csv):
            # Retrieve additional info from all_csv_rows if available
            info = all_csv_rows.get((group, el, id_val), (el, id_val, f"G{group}", "", ""))
            tip = info[3] if len(info) > 3 else "?"
            print(f"  G{group} EL={el}, ID={id_val}, Tip='{tip}'")
        print("="*60)
    else:
        print("\nAll CSV entries with target number were matched in Excel.")

    print("\nDone.")

if __name__ == "__main__":
    main()