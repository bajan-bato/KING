#!/usr/bin/env python3
"""
Scan the sort/ folder and generate an Excel report of all PDF files
with creation date only (YYYY-MM-DD) and alternating row colors per distinct date.
"""

import os
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

# ========== CONFIGURATION ==========
SORT_BASE = "sort"          # root folder containing G1/, G2/, etc.
OUTPUT_EXCEL = "sort_report.xlsx"
# ===================================

def get_file_date(filepath):
    """Return creation date as YYYY-MM-DD string (Windows: creation time, Unix: metadata change)."""
    # On Windows, getctime returns creation time; on Unix it returns metadata change time.
    # Since user is on Windows, this gives creation date.
    ctime = os.path.getctime(filepath)
    return datetime.fromtimestamp(ctime).strftime("%Y-%m-%d")

def main():
    if not os.path.isdir(SORT_BASE):
        print(f"Error: '{SORT_BASE}' folder not found.")
        return

    data = []
    # Walk through all subdirectories
    for root, dirs, files in os.walk(SORT_BASE):
        for file in files:
            if not file.lower().endswith(".pdf"):
                continue
            full_path = os.path.join(root, file)
            rel_path = os.path.relpath(root, SORT_BASE)
            parts = rel_path.split(os.sep)
            group = parts[0] if parts and parts[0] in ["G1", "G2"] else "Unknown"
            category = parts[1] if len(parts) > 1 else ""
            date_str = get_file_date(full_path)
            data.append({
                "File Name": file,
                "Full Path": full_path,
                "Group": group,
                "Category": category,
                "Date Created": date_str
            })

    if not data:
        print("No PDF files found.")
        return

    df = pd.DataFrame(data)
    # Sort by creation date (oldest first)
    df = df.sort_values(by="Date Created")

    # Write to Excel with conditional formatting for alternating colors per date
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # 1. Add a helper column "DateGroup" (hidden) that assigns a number to each unique date
        distinct_dates = df['Date Created'].unique()
        date_to_group = {date: i+1 for i, date in enumerate(distinct_dates)}  # 1,2,3,...
        helper_values = [date_to_group[d] for d in df['Date Created']]

        # Insert helper column after the Date Created column (Date Created is column E, so helper becomes F)
        helper_col_idx = df.columns.get_loc('Date Created') + 2   # +1 for 0‑based, +1 to put after Date
        worksheet.insert_cols(helper_col_idx)
        helper_cell = worksheet.cell(row=1, column=helper_col_idx)
        helper_cell.value = "DateGroup"
        for i, val in enumerate(helper_values, start=2):
            worksheet.cell(row=i, column=helper_col_idx, value=val)

        # 2. Define two light fills
        fill_even = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # light blue
        fill_odd  = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")  # light orange

        # 3. Apply conditional formatting to the entire data rows (columns A to E)
        helper_letter = openpyxl.utils.get_column_letter(helper_col_idx)
        last_data_col = openpyxl.utils.get_column_letter(df.columns.get_loc('Date Created') + 1)  # column E
        last_row = len(df) + 1
        range_str = f"A2:{last_data_col}{last_row}"

        # Even groups (DateGroup even)
        formula_even = f"=MOD(${helper_letter}2,2)=0"
        worksheet.conditional_formatting.add(range_str, FormulaRule(formula=[formula_even], fill=fill_even))
        # Odd groups (DateGroup odd)
        formula_odd  = f"=MOD(${helper_letter}2,2)=1"
        worksheet.conditional_formatting.add(range_str, FormulaRule(formula=[formula_odd], fill=fill_odd))

        # 4. Hide the helper column
        worksheet.column_dimensions[helper_letter].hidden = True

    print(f"Exported {len(df)} PDF files to {OUTPUT_EXCEL}")
    print("Rows are colored alternately per distinct creation date (light blue / light orange).")

if __name__ == "__main__":
    main()